import copy
import re
import os
import numpy as np
import openpyxl
import pandas as pd


def load_data(filename, sheet_name='data'):
    data_frame = pd.DataFrame(pd.read_excel(filename, sheet_name=sheet_name))
    data_set = data_frame.values
    return data_set


# Data acquisition
def get_data(path, sheet_name):
    data = pd.read_excel(path, sheet_name=sheet_name)
    columns = data.columns
    original_data = np.array(data)
    duplicate_data = copy.deepcopy(original_data[:, ])   #重新覆制了一个表格
    return original_data, duplicate_data, columns


def create_wb(path):
    if not os.path.exists(path):
        wb = openpyxl.Workbook()
        wb.save(filename=path)
        wb.close()
        print("Create workbook successfully!")


# 创建文件夹
def create_directory(path, location_index, method):
    reg_ex_1 = re.compile('/')
    list_of_tokens_1 = reg_ex_1.split(path)[location_index]  # data_90.xlsx
    # reg_ex_2 = re.compile('..')
    reg_ex_2 = re.compile('\.')
    list_of_tokens_2 = reg_ex_2.split(list_of_tokens_1)[0]  # data_90

    root_path = 'output_data/' + list_of_tokens_2
    statics_path = root_path + '/' + method
    figures_path = statics_path + '/figures'
    is_exists = os.path.exists(root_path)
    if is_exists:
        if not os.path.exists(statics_path):
            os.makedirs(statics_path)
            os.makedirs(figures_path)
        else:
            if not os.path.exists(figures_path):
                os.makedirs(figures_path)
    else:
        os.makedirs(root_path)
        os.makedirs(statics_path)
        os.makedirs(figures_path)
    return root_path, statics_path, figures_path


def save_to_excel_1d(data, column, wb_name, sheet_name, start_column, start_row):
    wb = openpyxl.load_workbook(filename=wb_name)
    try:
        sheet = wb[sheet_name]
    except:
        wb.create_sheet(title=sheet_name)
    finally:
        sheet = wb[sheet_name]
    _ = sheet.cell(row=1, column=start_column, value=str(column))
    for r in range(start_row, len(data) + start_row):
        _ = sheet.cell(row=r, column=start_column, value=str(data[r - start_row]))
    wb.save(filename=wb_name)
    wb.close()
    print("Save the file successfully!")


def save_to_excel_2d(data, columns, wb_name, sheet_name, start_column, start_row):
    wb = openpyxl.load_workbook(filename=wb_name)
    try:
        sheet = wb[sheet_name]
    except:
        wb.create_sheet(title=sheet_name)
    finally:
        sheet = wb[sheet_name]
    for field in range(start_column, len(columns) + start_column):
        _ = sheet.cell(row=1, column=field, value=str(columns[field - start_column]))
        for r in range(start_row, len(data) + start_row):
            _ = sheet.cell(row=r, column=field, value=str(data[r - start_row][field - start_column]))
    wb.save(filename=wb_name)
    wb.close()
    print("Save the file successfully!")


# Storing the data into the table ("num" is the started column number)
def save_to_excel_y(data, fields, sheet_name, wb_name, num):        # num = 从第几列开始存
    print("Writing.........")
    wb = openpyxl.load_workbook(filename=wb_name)
    try:
        sheet = wb[sheet_name]
    except:
        wb.create_sheet(title=sheet_name)
    sheet = wb[sheet_name]
    sheet.title = sheet_name
    for field in range(num, len(fields)+num):                                #用于写入表头
        _ = sheet.cell(row=1, column=field, value=str(fields[field-num]))
        for row1 in range(2, len(data) + 2):                                 #将表头后面的数据（列）存入excel中
             _ = sheet.cell(row=row1, column=num, value=str(data[row1 - 2]))
        wb.save(filename=wb_name)
        print("Successfully save!")